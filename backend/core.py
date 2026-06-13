from __future__ import annotations

import csv
import base64
import hashlib
import io
import math
import secrets
from datetime import date, datetime, timedelta
from typing import Any


TAG_LABELS = {
    "zh": {
        "highSalt": "高盐",
        "highCarb": "高碳水",
        "diningOut": "外食",
        "training": "训练日",
        "lateNight": "熬夜",
        "stress": "压力大",
    },
    "en": {
        "highSalt": "high salt",
        "highCarb": "high carb",
        "diningOut": "dining out",
        "training": "training day",
        "lateNight": "late night",
        "stress": "stress",
    },
}


def utc_now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat() + "Z"


def session_expires_at(now: str | None = None, days: int = 30) -> str:
    base = parse_iso_utc(now) if now else datetime.utcnow().replace(microsecond=0)
    return (base + timedelta(days=days)).replace(microsecond=0).isoformat() + "Z"


def parse_iso_utc(value: str) -> datetime:
    return datetime.fromisoformat(value.rstrip("Z"))


def create_session_token() -> str:
    return secrets.token_urlsafe(32)


def hash_session_token(token: str) -> str:
    if not token:
        raise ValueError("token is required")
    return hashlib.sha256(token.encode("utf-8")).hexdigest()


def normalize_account(account: str) -> str:
    normalized = (account or "").strip().lower()
    if not normalized:
        raise ValueError("account is required")
    return normalized


def create_user(account: str, created_at: str | None = None) -> dict[str, Any]:
    normalized = normalize_account(account)
    created = created_at or utc_now_iso()
    digest = hashlib.sha1(f"{normalized}:{created}".encode("utf-8")).hexdigest()[:12]
    return {
        "id": f"user-{normalized.replace('@', '-at-')}-{digest}",
        "account": normalized,
        "displayName": (account or "").strip(),
        "language": "zh",
        "targetWeightKg": 68,
        "createdAt": created,
    }


def hash_password(password: str, salt: str | None = None) -> str:
    if not password:
        raise ValueError("password is required")
    actual_salt = salt or secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), actual_salt.encode("utf-8"), 260000)
    return f"pbkdf2_sha256${actual_salt}${base64.b64encode(digest).decode('ascii')}"


def verify_password(password: str, password_hash: str | None) -> bool:
    if not password or not password_hash:
        return False
    try:
        algorithm, salt, expected = password_hash.split("$", 2)
    except ValueError:
        return False
    if algorithm != "pbkdf2_sha256":
        return False
    candidate = hash_password(password, salt).split("$", 2)[2]
    return secrets.compare_digest(candidate, expected)


def is_number(value: Any) -> bool:
    if value is None or value == "":
        return True
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(number)


def to_number(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    if not math.isfinite(number):
        return default
    return number


def validate_daily_record(record: dict[str, Any]) -> list[str]:
    errors: list[str] = []
    if not record.get("date"):
        errors.append("date")
    for field in ("weightKg", "sleepHours", "exerciseCalories", "caloriesIn", "proteinG", "carbsG", "fatG"):
        if not is_number(record.get(field)):
            errors.append(field)
    return errors


def clean_daily_record(record: dict[str, Any]) -> dict[str, Any]:
    tags = record.get("tags") or []
    if isinstance(tags, str):
        tags = split_tags(tags)
    exercise_items = record.get("exerciseItems") or []
    if isinstance(exercise_items, str):
        exercise_items = [{"name": exercise_items, "durationMinutes": None, "caloriesBurned": to_number(record.get("exerciseCalories"), 0)}]
    return {
        "id": record.get("id"),
        "userId": record.get("userId"),
        "date": record.get("date"),
        "weightKg": to_number(record.get("weightKg")),
        "foodText": record.get("foodText", ""),
        "caloriesIn": to_number(record.get("caloriesIn")),
        "proteinG": to_number(record.get("proteinG")),
        "carbsG": to_number(record.get("carbsG")),
        "fatG": to_number(record.get("fatG")),
        "foodItems": record.get("foodItems") or [],
        "exerciseItems": exercise_items,
        "exerciseCalories": to_number(record.get("exerciseCalories"), 0) or 0,
        "sleepHours": to_number(record.get("sleepHours")),
        "tags": tags,
        "note": record.get("note", ""),
        "createdAt": record.get("createdAt"),
        "updatedAt": record.get("updatedAt"),
    }


def upsert_daily_record(records: list[dict[str, Any]], record: dict[str, Any], now: str | None = None) -> list[dict[str, Any]]:
    errors = validate_daily_record(record)
    if errors:
        raise ValueError(",".join(errors))
    timestamp = now or utc_now_iso()
    cleaned = clean_daily_record(record)
    cleaned["id"] = cleaned.get("id") or f"record-{cleaned['userId']}-{cleaned['date']}"
    cleaned["createdAt"] = cleaned.get("createdAt") or timestamp
    cleaned["updatedAt"] = timestamp

    result: list[dict[str, Any]] = []
    replaced = False
    for existing in records:
        same_day = existing.get("userId") == cleaned.get("userId") and existing.get("date") == cleaned.get("date")
        if same_day:
            merged = {**existing, **cleaned, "createdAt": existing.get("createdAt") or cleaned["createdAt"], "updatedAt": timestamp}
            result.append(merged)
            replaced = True
        else:
            result.append(existing)
    if not replaced:
        result.append(cleaned)
    return sort_records(result)


def records_for_user(records: list[dict[str, Any]], user_id: str) -> list[dict[str, Any]]:
    return sort_records([record for record in records if record.get("userId") == user_id])


def sort_records(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return sorted(records, key=lambda item: item.get("date") or "")


def compute_moving_average(values: list[float | None], window: int) -> list[float | None]:
    output: list[float | None] = []
    for index in range(len(values)):
        chunk = values[index - window + 1 : index + 1]
        if len(chunk) < window or any(value is None for value in chunk):
            output.append(None)
        else:
            output.append(round(sum(float(value) for value in chunk) / window, 2))
    return output


def build_chart_series(records: list[dict[str, Any]]) -> dict[str, Any]:
    ordered = sort_records(records)
    weights = [to_number(record.get("weightKg")) for record in ordered]
    return {
        "dates": [record.get("date") for record in ordered],
        "weights": weights,
        "ma7": compute_moving_average(weights, 7),
        "ma14": compute_moving_average(weights, 14),
        "exerciseCalories": [to_number(record.get("exerciseCalories"), 0) or 0 for record in ordered],
        "sleepHours": [to_number(record.get("sleepHours")) for record in ordered],
        "caloriesIn": [to_number(record.get("caloriesIn")) for record in ordered],
    }


def summarize_changes(records: list[dict[str, Any]]) -> dict[str, float | None]:
    weighted = [record for record in sort_records(records) if to_number(record.get("weightKg")) is not None]

    def change(days: int) -> float | None:
        if len(weighted) <= 1:
            return None
        recent = weighted[-days:] if len(weighted) >= days else weighted
        if len(recent) < 2:
            return None
        return round(float(recent[-1]["weightKg"]) - float(recent[0]["weightKg"]), 2)

    return {"sevenDayKg": change(7), "thirtyDayKg": change(30)}


def summarize_goal_progress(records: list[dict[str, Any]], target_weight_kg: float | None) -> dict[str, Any]:
    weighted = [record for record in sort_records(records) if to_number(record.get("weightKg")) is not None]
    if not weighted or target_weight_kg is None:
        return {"currentWeightKg": None, "targetWeightKg": target_weight_kg, "distanceKg": None, "completionPercent": 0, "estimatedDate": None}
    start = float(weighted[0]["weightKg"])
    current = float(weighted[-1]["weightKg"])
    total = abs(start - float(target_weight_kg))
    moved = abs(start - current)
    completion = 0 if total == 0 else max(0, min(100, round((moved / total) * 100)))
    distance = round(abs(current - float(target_weight_kg)), 2)
    changes = summarize_changes(weighted)
    daily_delta = None
    if len(weighted) > 1:
        daily_delta = (current - start) / (len(weighted) - 1)
    estimated = None
    if daily_delta and ((target_weight_kg - current) / daily_delta) > 0:
        days = min(365, max(1, round((target_weight_kg - current) / daily_delta)))
        estimated = (parse_date(weighted[-1]["date"]) + timedelta(days=days)).isoformat()
    return {
        "currentWeightKg": round(current, 2),
        "targetWeightKg": target_weight_kg,
        "distanceKg": distance,
        "completionPercent": completion,
        "estimatedDate": estimated,
        "changes": changes,
    }


def generate_prediction(records: list[dict[str, Any]], language: str = "zh") -> list[dict[str, Any]]:
    weighted = [record for record in sort_records(records) if to_number(record.get("weightKg")) is not None]
    if not weighted:
        return []
    recent = weighted[-7:]
    last = recent[-1]
    last_weight = float(last["weightKg"])
    trend = 0.0
    if len(recent) > 1:
        trend = (float(recent[-1]["weightKg"]) - float(recent[0]["weightKg"])) / (len(recent) - 1)

    tags = set(last.get("tags") or [])
    adjustment = 0.0
    factor_keys: list[str] = []
    if "highSalt" in tags:
        adjustment += 0.12
        factor_keys.append("highSalt")
    if "highCarb" in tags:
        adjustment += 0.1
        factor_keys.append("highCarb")
    if "diningOut" in tags:
        adjustment += 0.06
        factor_keys.append("diningOut")
    if to_number(last.get("sleepHours")) is not None and float(last["sleepHours"]) < 6:
        adjustment += 0.1
        factor_keys.append("sleepShort")
    if (to_number(last.get("exerciseCalories"), 0) or 0) >= 300:
        adjustment -= 0.06
        factor_keys.append("exercise")

    confidence = "medium" if len(weighted) >= 7 else "low"
    width = 0.18 if confidence == "medium" else 0.28
    labels = prediction_labels(language)
    factors = [labels["factors"].get(key, key) for key in factor_keys] or [labels["factors"]["trend"]]
    suggestion = labels["suggestion_with_factors"] if factor_keys else labels["suggestion_trend"]

    output: list[dict[str, Any]] = []
    source = parse_date(last["date"])
    for offset in (1, 2):
        center = last_weight + (trend * offset) + (adjustment * (1.0 if offset == 1 else 0.65))
        output.append({
            "sourceDate": last["date"],
            "targetDate": (source + timedelta(days=offset)).isoformat(),
            "minWeightKg": round(center - width, 2),
            "maxWeightKg": round(center + width, 2),
            "confidence": confidence,
            "factors": factors,
            "suggestion": suggestion,
        })
    return output


def prediction_labels(language: str) -> dict[str, Any]:
    if language == "en":
        return {
            "factors": {
                "highSalt": "high salt may raise short-term water weight",
                "highCarb": "high carb may raise glycogen and water weight",
                "diningOut": "dining out can increase sodium uncertainty",
                "sleepShort": "short sleep may increase fluctuation risk",
                "exercise": "exercise expenditure can partially offset intake",
                "trend": "recent weight trend",
            },
            "suggestion_with_factors": "Treat this as a fluctuation range: salt, carbs, sleep, and exercise mostly affect short-term water weight, not instant fat gain or loss.",
            "suggestion_trend": "Prediction is based mostly on recent trend. Keep recording for a tighter and more reliable range.",
        }
    return {
        "factors": {
            "highSalt": "高盐可能带来短期水重上浮",
            "highCarb": "高碳水可能增加糖原和水重",
            "diningOut": "外食会增加钠摄入不确定性",
            "sleepShort": "睡眠不足可能提高波动风险",
            "exercise": "运动消耗可部分抵消摄入",
            "trend": "近期体重趋势",
        },
        "suggestion_with_factors": "把它当作波动区间看待：盐、碳水、睡眠和运动主要影响短期水重，不代表即时脂肪增减。",
        "suggestion_trend": "当前预测主要基于近期趋势。继续记录后，区间会更稳定可信。",
    }


KG_WEIGHT_HEADERS = {
    "weightkg",
    "kg",
    "kilogram",
    "kilograms",
    "weight(kg)",
    "weight（kg）",
    "weight公斤",
    "体重kg",
    "体重(kg)",
    "体重（kg）",
    "体重公斤",
    "体重(公斤)",
    "体重（公斤）",
    "公斤",
    "千克",
}
JIN_WEIGHT_HEADERS = {
    "weightjin",
    "jin",
    "weight(斤)",
    "weight（斤）",
    "weight斤",
    "体重jin",
    "体重斤",
    "体重(斤)",
    "体重（斤）",
    "斤",
}
GENERIC_WEIGHT_HEADERS = {"weight", "体重"}


def parse_csv_records(text: str, weight_unit: str = "kg") -> list[dict[str, Any]]:
    stream = io.StringIO(text.strip())
    rows: list[dict[str, Any]] = []
    for raw in csv.DictReader(stream):
        row = {key.strip(): (value or "").strip() for key, value in raw.items() if key}
        rows.append(clean_daily_record({
            "date": row.get("date") or row.get("日期"),
            "weightKg": import_weight_kg(row, weight_unit),
            "foodText": row.get("foodText") or row.get("food") or row.get("饮食") or "",
            "caloriesIn": row.get("caloriesIn") or row.get("calories") or row.get("热量"),
            "proteinG": row.get("proteinG") or row.get("protein") or row.get("蛋白质"),
            "carbsG": row.get("carbsG") or row.get("carbs") or row.get("碳水"),
            "fatG": row.get("fatG") or row.get("fat") or row.get("脂肪"),
            "exerciseCalories": row.get("exerciseCalories") or row.get("exerciseKcal") or row.get("运动消耗"),
            "sleepHours": row.get("sleepHours") or row.get("sleep") or row.get("睡眠"),
            "tags": row.get("tags") or row.get("标签") or "",
            "note": row.get("note") or row.get("备注") or "",
        }))
    return rows


def parse_excel_records(content: bytes, weight_unit: str = "kg") -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to parse Excel files") from exc

    workbook = load_workbook(io.BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [cell_text(value) for value in rows[0]]
    parsed: list[dict[str, Any]] = []
    for values in rows[1:]:
        raw = {headers[index]: cell_text(value) for index, value in enumerate(values) if index < len(headers) and headers[index]}
        if not any(raw.values()):
            continue
        parsed.append(clean_daily_record({
            "date": raw.get("date") or raw.get("日期"),
            "weightKg": import_weight_kg(raw, weight_unit),
            "foodText": raw.get("foodText") or raw.get("food") or raw.get("饮食") or "",
            "caloriesIn": raw.get("caloriesIn") or raw.get("calories") or raw.get("热量"),
            "proteinG": raw.get("proteinG") or raw.get("protein") or raw.get("蛋白质"),
            "carbsG": raw.get("carbsG") or raw.get("carbs") or raw.get("碳水"),
            "fatG": raw.get("fatG") or raw.get("fat") or raw.get("脂肪"),
            "exerciseCalories": raw.get("exerciseCalories") or raw.get("exerciseKcal") or raw.get("运动消耗"),
            "sleepHours": raw.get("sleepHours") or raw.get("sleep") or raw.get("睡眠"),
            "tags": raw.get("tags") or raw.get("标签") or "",
            "note": raw.get("note") or raw.get("备注") or "",
        }))
    return parsed


def import_weight_kg(row: dict[str, Any], fallback_unit: str = "kg") -> float | None:
    for key, value in row.items():
        if normalize_header(key) in KG_WEIGHT_HEADERS:
            return to_number(value)
    for key, value in row.items():
        if normalize_header(key) in JIN_WEIGHT_HEADERS:
            return imported_weight_to_kg(value, "jin")
    for key, value in row.items():
        if normalize_header(key) in GENERIC_WEIGHT_HEADERS:
            return imported_weight_to_kg(value, fallback_unit)
    return None


def imported_weight_to_kg(value: Any, unit: str = "kg") -> float | None:
    number = to_number(value)
    if number is None:
        return None
    if normalize_weight_unit(unit) == "jin":
        return round(number / 2, 3)
    return number


def normalize_weight_unit(unit: Any) -> str:
    text = str(unit or "").strip().lower()
    return "jin" if text in {"jin", "斤"} else "kg"


def normalize_header(value: Any) -> str:
    return cell_text(value).replace(" ", "").replace("_", "").lower()


def generate_sample_excel() -> bytes:
    try:
        from openpyxl import Workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to generate Excel files") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "between-points"
    sheet.append(["date", "weightKg", "foodText", "exerciseCalories", "sleepHours", "note"])
    sheet.append(["2026-06-01", 72.6, "早餐燕麦鸡蛋，午餐牛肉饭", 180, 7, "示例数据"])
    sheet.append(["2026-06-02", 72.4, "家常餐", 260, 6.5, "训练日"])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def split_tags(value: str) -> list[str]:
    normalized = value.replace("|", ";").replace(",", ";").replace("，", ";")
    return [tag.strip() for tag in normalized.split(";") if tag.strip()]


def apply_import_rows(records: list[dict[str, Any]], user_id: str, rows: list[dict[str, Any]], now: str | None = None) -> list[dict[str, Any]]:
    result = records[:]
    timestamp = now or utc_now_iso()
    for row in rows:
        result = upsert_daily_record(result, {**row, "userId": user_id}, timestamp)
    return result


def analyze_records(records: list[dict[str, Any]], predictions: list[dict[str, Any]] | None = None) -> dict[str, Any]:
    ordered = sort_records(records)
    with_weights = [record for record in ordered if to_number(record.get("weightKg")) is not None]
    tag_totals: dict[str, list[float]] = {}
    for index, record in enumerate(with_weights[:-1]):
        next_record = with_weights[index + 1]
        delta = round(float(next_record["weightKg"]) - float(record["weightKg"]), 2)
        for tag in record.get("tags") or []:
            tag_totals.setdefault(tag, []).append(delta)
    tag_impact = [
        {"tag": tag, "averageDeltaKg": round(sum(values) / len(values), 2), "count": len(values)}
        for tag, values in sorted(tag_totals.items())
    ]
    calories = [to_number(record.get("caloriesIn")) for record in ordered if to_number(record.get("caloriesIn")) is not None]
    sleep = [to_number(record.get("sleepHours")) for record in ordered if to_number(record.get("sleepHours")) is not None]
    return {
        "recordCount": len(ordered),
        "enoughData": len(ordered) >= 7,
        "averageCalories": round(sum(calories) / len(calories)) if calories else None,
        "averageSleepHours": round(sum(float(value) for value in sleep) / len(sleep), 1) if sleep else None,
        "tagImpact": tag_impact,
        "predictionAccuracy": predictions or [],
    }


def parse_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()
